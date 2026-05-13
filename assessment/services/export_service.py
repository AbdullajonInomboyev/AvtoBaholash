"""
AvtoBaholash Export Services
- PDF: topshiriqni PDF holatida yuklab olish
- Excel: baholarni Excel holatida yuklab olish  
- Telegram: bildirishnomalarni Telegram botga yuborish
"""
import io
import logging
import json
from django.http import HttpResponse

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────
def generate_assignment_pdf(assignment):
    """
    Topshiriqni PDF holatida qaytaradi.
    O'qituvchi offline tarqatish uchun yuklab olishi mumkin.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return _fallback_pdf_response(assignment)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    primary = colors.HexColor('#00236f')
    secondary = colors.HexColor('#0058be')
    light_bg = colors.HexColor('#f0f4ff')

    title_style = ParagraphStyle(
        'Title', parent=styles['Normal'],
        fontSize=16, fontName='Helvetica-Bold',
        textColor=primary, spaceAfter=6, alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#6b7280'),
        spaceAfter=4, alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'Heading', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold',
        textColor=primary, spaceAfter=6
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['Normal'],
        fontSize=10, spaceAfter=4, leading=14
    )
    question_style = ParagraphStyle(
        'Question', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold',
        spaceAfter=3, leading=13
    )
    option_style = ParagraphStyle(
        'Option', parent=styles['Normal'],
        fontSize=10, leftIndent=20, spaceAfter=2
    )

    story = []

    # Header
    story.append(Paragraph("AvtoBaholash", ParagraphStyle('Brand', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(assignment.title, title_style))
    story.append(Paragraph(
        f"{assignment.subject.name} · {assignment.get_assignment_type_display()} · "
        f"Muddat: {assignment.deadline.strftime('%d.%m.%Y %H:%M')}",
        subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=primary))
    story.append(Spacer(1, 0.4*cm))

    # Meta table
    meta_data = [
        ['Fan:', assignment.subject.name, "O'qituvchi:", assignment.teacher.full_name],
        ['Muddat:', assignment.deadline.strftime('%d.%m.%Y %H:%M'), 'Max ball:', f"{assignment.max_score:.0f}"],
    ]
    if assignment.assignment_type == 'test':
        meta_data.append(['Davomiyligi:', f"{assignment.duration_minutes} daqiqa", 'Savollar:', str(assignment.questions.count())])

    meta_table = Table(meta_data, colWidths=[3*cm, 6.5*cm, 3*cm, 4.5*cm])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light_bg),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), primary),
        ('TEXTCOLOR', (2, 0), (2, -1), primary),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [light_bg, colors.white]),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2ed')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2ed')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.4*cm))

    # Description
    if assignment.description:
        story.append(Paragraph("Topshiriq:", heading_style))
        story.append(Paragraph(assignment.description, body_style))
        story.append(Spacer(1, 0.3*cm))

    if assignment.instructions:
        story.append(Paragraph("Ko'rsatmalar:", heading_style))
        story.append(Paragraph(assignment.instructions, body_style))
        story.append(Spacer(1, 0.3*cm))

    # Questions (for tests)
    if assignment.assignment_type == 'test':
        questions = assignment.questions.order_by('order')
        if questions.exists():
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dee2ed')))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph("Savollar", heading_style))
            story.append(Spacer(1, 0.2*cm))

            for i, q in enumerate(questions, 1):
                story.append(Paragraph(f"{i}. {q.text}", question_style))
                if q.image:
                    pass  # Image handling
                for key, text in q.get_options():
                    story.append(Paragraph(f"   {key}) {text}", option_style))
                # Answer line for offline use
                story.append(Paragraph("   Javob: _______", option_style))
                story.append(Spacer(1, 0.2*cm))

    # Written answer space
    elif assignment.assignment_type == 'written':
        story.append(Paragraph("Javob:", heading_style))
        for _ in range(15):
            story.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#dee2ed')))
            story.append(Spacer(1, 0.55*cm))

    # Footer
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dee2ed')))
    story.append(Paragraph(
        "Talaba: _________________________ | Guruh: ________ | Imzo: _________",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9,
                       textColor=colors.HexColor('#9ca3af'), spaceAfter=4)
    ))
    story.append(Paragraph(
        "AvtoBaholash · AI yordamida baholash tizimi · edulens.uz",
        ParagraphStyle('FooterBrand', parent=styles['Normal'], fontSize=8,
                       textColor=colors.HexColor('#c5c5d3'), alignment=TA_CENTER)
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _fallback_pdf_response(assignment):
    """ReportLab mavjud bo'lmasa oddiy matn qaytaradi"""
    lines = [
        f"AvtoBaholash — {assignment.title}",
        f"Fan: {assignment.subject.name}",
        f"Muddat: {assignment.deadline.strftime('%d.%m.%Y %H:%M')}",
        f"O'qituvchi: {assignment.teacher.full_name}",
        "",
        assignment.description or "",
    ]
    if assignment.assignment_type == 'test':
        for i, q in enumerate(assignment.questions.order_by('order'), 1):
            lines.append(f"\n{i}. {q.text}")
            for key, text in q.get_options():
                lines.append(f"   {key}) {text}")
    return "\n".join(lines).encode('utf-8')


def assignment_pdf_response(assignment):
    """Django HttpResponse qaytaradi"""
    try:
        pdf_bytes = generate_assignment_pdf(assignment)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        safe_title = assignment.title[:40].replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{safe_title}.pdf"'
        return response
    except Exception as e:
        logger.error(f"PDF xato: {e}")
        response = HttpResponse(f"PDF yaratishda xato: {e}", content_type='text/plain')
        return response


# ─────────────────────────────────────────
# EXCEL EXPORT (kengaytirilgan)
# ─────────────────────────────────────────
def export_gradebook_xlsx(assignment, submissions):
    """Baholash jurnalini Excel formatida qaytaradi"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baholar jurnali"

    # Styles
    header_fill = PatternFill("solid", fgColor="00236F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=13, color="00236F")
    center_aln  = Alignment(horizontal="center", vertical="center")
    border      = Border(
        left=Side(style='thin', color='DEE2ED'),
        right=Side(style='thin', color='DEE2ED'),
        top=Side(style='thin', color='DEE2ED'),
        bottom=Side(style='thin', color='DEE2ED')
    )

    # Title rows
    ws.merge_cells('A1:H1')
    ws['A1'] = f"AvtoBaholash — Baholash Jurnali"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_aln

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Topshiriq: {assignment.title} | Fan: {assignment.subject.name} | O'qituvchi: {assignment.teacher.full_name}"
    ws['A2'].font = Font(size=10, color="6B7280")
    ws['A2'].alignment = center_aln

    ws.merge_cells('A3:H3')
    ws['A3'] = f"Muddat: {assignment.deadline.strftime('%d.%m.%Y %H:%M')} | Jami: {submissions.count()} ta javob"
    ws['A3'].font = Font(size=9, color="9CA3AF")
    ws['A3'].alignment = center_aln

    # Headers row 5
    headers = ["#", "Talaba F.I.O", "Guruh", "ID", "AI ball", "Yakuniy ball", "Baho", "Izoh", "Topshirilgan", "Holat"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aln
        cell.border = border

    # Column widths
    col_widths = [5, 28, 10, 12, 10, 12, 8, 30, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[5].height = 22

    # Data rows
    grade_colors = {'green': 'D1FAE5', 'yellow': 'FEF3C7', 'red': 'FEE2E2', 'gray': 'F3F4F6'}

    for i, sub in enumerate(submissions, 1):
        row = 5 + i
        data = [
            i,
            sub.student.full_name,
            str(sub.student.group) if sub.student.group else "—",
            sub.student.student_id or "—",
            sub.ai_score or "—",
            sub.final_score or "—",
            sub.grade_letter,
            sub.teacher_note or "",
            sub.submitted_at.strftime('%d.%m.%Y %H:%M'),
            sub.get_status_display(),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 8))
            if col == 7 and sub.grade_color in grade_colors:
                cell.fill = PatternFill("solid", fgColor=grade_colors[sub.grade_color])
                cell.font = Font(bold=True)
            if i % 2 == 0 and col != 7:
                cell.fill = PatternFill("solid", fgColor="F8FAFF")

    # Freeze panes
    ws.freeze_panes = 'A6'

    # Stats sheet
    ws2 = wb.create_sheet("Statistika")
    ws2['A1'] = "Ko'rsatkich"
    ws2['B1'] = "Qiymat"
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    
    graded_subs = [s for s in submissions if s.final_score is not None]
    scores = [s.final_score for s in graded_subs]
    stats = [
        ("Jami javoblar", submissions.count()),
        ("Baholangan", len(graded_subs)),
        ("O'rtacha ball", f"{sum(scores)/len(scores):.1f}" if scores else "—"),
        ("Eng yuqori", f"{max(scores):.1f}" if scores else "—"),
        ("Eng past", f"{min(scores):.1f}" if scores else "—"),
        ("90+ (A+)", sum(1 for s in scores if s >= 90)),
        ("80-89 (A)", sum(1 for s in scores if 80 <= s < 90)),
        ("70-79 (B)", sum(1 for s in scores if 70 <= s < 80)),
        ("60-69 (C)", sum(1 for s in scores if 60 <= s < 70)),
        ("50-59 (D)", sum(1 for s in scores if 50 <= s < 60)),
        ("50 dan past (F)", sum(1 for s in scores if s < 50)),
    ]
    for r, (k, v) in enumerate(stats, 2):
        ws2.cell(row=r, column=1, value=k).font = Font(color="374151")
        ws2.cell(row=r, column=2, value=v).font = Font(bold=True, color="00236F")
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─────────────────────────────────────────
# TELEGRAM WEBHOOK
# ─────────────────────────────────────────
def send_telegram_message(bot_token, chat_id, text, parse_mode='HTML'):
    """
    Telegram botiga xabar yuborish.
    Bot token va chat_id ni .env faylda sozlang:
      TELEGRAM_BOT_TOKEN=...
      TELEGRAM_CHAT_ID=...
    """
    if not bot_token or not chat_id:
        return False
    try:
        import urllib.request
        import urllib.parse
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        data = urllib.parse.urlencode({
            'chat_id': chat_id,
            'text': text,
            'parse_mode': parse_mode,
        }).encode()
        req = urllib.request.Request(url, data=data, method='POST')
        with urllib.request.urlopen(req, timeout=5) as resp:
            return resp.status == 200
    except Exception as e:
        logger.warning(f"Telegram xato: {e}")
        return False


def notify_telegram(user, title, message):
    """
    Foydalanuvchiga Telegram xabari yuborish.
    User modelida telegram_chat_id bo'lsa ishlatiladi.
    """
    from django.conf import settings
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
    
    # Individual user notification
    chat_id = getattr(user, 'telegram_chat_id', None)
    if chat_id and bot_token:
        text = f"🎓 <b>AvtoBaholash</b>\n\n<b>{title}</b>\n{message}"
        send_telegram_message(bot_token, chat_id, text)

    # Department-wide channel
    dept = getattr(user, 'department', None)
    if dept:
        channel_id = getattr(dept, 'telegram_channel_id', None)
        if channel_id and bot_token:
            text = f"🎓 <b>AvtoBaholash</b>\n\n<b>{title}</b>\n{message}"
            send_telegram_message(bot_token, channel_id, text)
